"""
 Created by waldo on 2/10/17
"""

import csv
import os
import re
import mechanize
import logging
import time
import pdb
from bs4 import BeautifulSoup
from enum import Enum
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

__author__ = "waldo"
__project__ = "prj_socialdata_backend"

logger = logging.getLogger(__name__)


class SocialMedia(Enum):
    ALL = 0
    FACEBOOK = 1
    LINKEDIN = 2


class FileFormat(Enum):
    ALL = 0
    CSV = 1
    EXCEL = 2


class Crawler(object):
    def __init__(self):
        self._current_session = None

    def open_session(self, social_media, user_name, password):
        if social_media == "FACEBOOK":
            self._current_session = SocialMedium(user_name, password)  # TODO: Change name SocialMedium man...

    def get_contacts_file(self, profile_id, file_format, file_path,
                          progress_callback=None):  # TODO Return file handle instead of path
        max_percentage = 90

        total_num_contacts = self._current_session.num_contacts(profile_id)

        logger.debug("Total number of contacts found: {0}.".format(total_num_contacts))

        for contact_list in self._current_session.get_contact_list(profile_id):
            progress = int(((len(contact_list) - 1) / float(
                total_num_contacts)) * max_percentage)  # TODO: we still have to make a file so this should not be 100% already
            logger.debug("Calculating progress...{0}%.".format(progress))

            try:
                progress_callback(progress)
            except:
                # Command line mode will get us here
                pass

        file_path = self._list_to_file(contact_list[:1], contact_list[1:], file_format, file_path)

        try:
            progress_callback(100)
        except:
            # Command line mode will get us here
            print "Calculating progress...{0}%.".format(progress)

        return file_path

    def get_mutual_contacts_file(self, profile_id1, profile_id2, file_format, file_path,
                                 progress_callback=None):  # TODO Return file handle instead of path
        # contact_list = self._current_session.get_mutual_contact_list(profile_id1, profile_id2)
        num_contacts1 = self._current_session.num_contacts(profile_id1)

        try:
            progress_callback(5)
        except:
            # Command line mode will get us here
            pass

        num_contacts2 = self._current_session.num_contacts(profile_id2)

        try:
            base_progress = 10
            progress_callback(base_progress)
        except:
            # Command line mode will get us here
            pass

        total_num_contacts = num_contacts1 + num_contacts2

        max_percentage = 80
        current_progress = base_progress
        for contact_list in self._current_session.get_mutual_contact_list(profile_id1, profile_id2):
            progress = base_progress + int(
                ((len(contact_list) - 2) / float(total_num_contacts)) * max_percentage)  # -2?

            if progress > current_progress:
                current_progress = progress

            logger.debug("Calculating progress...{0}%.".format(current_progress))

            try:
                progress_callback(current_progress)
            except:
                # Command line mode will get us here
                print "Calculating progress...{0}%.".format(current_progress)

        file_path = self._list_to_file(contact_list[:2], contact_list[2:], file_format, file_path)

        try:
            progress_callback(100)
        except:
            # Command line mode will get us here
            print "Calculating progress...{0}%.".format(100)

        return file_path

    def close_session(self):
        self._current_session.logout()

    @staticmethod
    def _list_to_file(profile_list, contact_list, file_format, file_path):
        if file_format == "CSV":
            logger.debug("Converting contact list to CSV file.")

            with open(file_path, 'wb') as csvfile:
                file_path = os.path.realpath(csvfile.name)
                csvfile.write(u'\ufeff'.encode('utf8'))
                writer = csv.DictWriter(csvfile, ["name", "profile_id", "uri"])

                # Write the data belonging to the profiles being researched
                for contact in profile_list:
                    writer.writerow({k: v.encode('utf8') for k, v in contact.items()})

                # Write an empty row
                writer.writerow({"name": "", "profile_id": "", "uri": ""})

                # Write the actual contacts
                writer.writeheader()
                if not contact_list:
                    writer.writerow({"name": "No contacts to show.", "profile_id": "", "uri": ""})
                else:
                    for contact in contact_list:
                        writer.writerow({k: v.encode('utf8') for k, v in contact.items()})

                logger.debug("File created.")

            logger.debug("File may or may not be created. Check previous log messages.")
        elif file_format == "EXCEL":
            logger.debug("Converting contact list to Excel file.")

            contact_book = ContactWorkbook()

            # Write the data belonging to the profiles being researched
            for contact in profile_list:
                contact_book.write_row([contact["name"], contact["profile_id"], contact["uri"]], bold=True)

            # Write an empty row
            contact_book.write_row(["", "", ""], bold=False)

            # Write the actual contacts
            contact_book.write_contacts(["name", "profile_id", "uri"], contact_list)
            file_path = os.path.realpath(file_path)
            contact_book.save(file_path)

            logger.debug("File created.")

        return file_path


class ContactWorkbook(object):
    def __init__(self):
        self.workbook = Workbook()
        self.current_sheet = self.workbook.active
        self.num_rows = 0
        self.num_cols = 0

    def load(self, filename):
        self.workbook = load_workbook(filename=filename, data_only=True)
        self.workbook.guess_types = True
        self.current_sheet = self.workbook.active
        self.num_rows = self.current_sheet.max_row
        self.num_cols = self.current_sheet.max_column

    def write_row(self, data_list, bold=False):
        font = Font(name="Calibri", size=11, bold=bold)

        def styled_cells(data):
            for col_nr, elem in enumerate(data):
                col_letter = get_column_letter(col_nr + 1)
                cell = Cell(self.current_sheet, column=col_letter, row=1, value=elem)
                cell.font = font
                yield cell

        self.current_sheet.append(styled_cells(data_list))

    def write_contacts(self, headers, contact_list):
        # Print the headers
        self.current_sheet.append(headers)

        if not contact_list:
            self.current_sheet.append(["No contacts to show."])
        else:
            # For each contact, print a row in the workbook
            for contact_dict in contact_list:
                self.current_sheet.append([contact_dict["name"], contact_dict["profile_id"], contact_dict["uri"]])

    def save(self, filepath):
        self.workbook.save(filepath)


class SocialMedium(object):
    login_url = 'https://www.facebook.com/login.php'
    user_agent = 'Mozilla/5.0 (X11; U; Linux i686; en-US) AppleWebKit/534.7 (KHTML, like Gecko) ' \
                 'Chrome/7.0.517.41 Safari/534.7'

    def __init__(self, user_name, password):
        self.name = ""
        self.url = "https://www.facebook.com/{0}".format(user_name)
        self.user_name = user_name
        self.password = password
        self.browser = mechanize.Browser()
        after_login_page = self.login()
        self._extract_login_data(after_login_page)

    def login(self):
        self.browser.set_handle_robots(False)
        cookies = mechanize.CookieJar()
        self.browser.set_cookiejar(cookies)
        self.browser.addheaders = [('User-agent', self.user_agent)]
        self.browser.set_handle_refresh(False)
        self.browser.open(self.login_url).read()  # Login page
        self.browser.select_form(nr=0)  # This is login-password form -> nr = number = 0
        self.browser.form['email'] = self.user_name
        self.browser.form['pass'] = self.password
        response = self.browser.submit()

        return response.read()

    def logout(self):
        self.browser.close()

    def num_contacts(self, profile_name):
        rolodex = Rolodex(self.browser, self.login_id, self.start_time,
                          profile_name)  # TODO: Write Singleton wrapper (Session) for browser instead of passing around. You can put multiple fields in Session, so you don't have to pass them around either
        return rolodex.num_entries()

    def get_contact_list(self, profile_name):
        contact_list = []

        logger.debug("Building contact list.")

        rolodex = Rolodex(self.browser, self.login_id, self.start_time,
                          profile_name)  # TODO: Write Singleton wrapper (Session) for browser instead of passing around. You can put multiple fields in Session, so you don't have to pass them around either

        logger.debug("Browsing contact list.")

        for page_nr, page_list in enumerate(rolodex):
            logger.debug("Page {0}.".format(page_nr + 1))
            contact_list.extend(page_list)
            yield contact_list

    def get_mutual_contact_list(self, profile_id1, profile_id2):
        # Get two dictionaries representing the contact lists of both accounts
        logger.debug("Building contact list for ID1.")
        for contacts_list_id1 in self.get_contact_list(profile_id1):
            yield contacts_list_id1

        logger.debug("Building contact list for ID2.")
        for contacts_list_id2 in self.get_contact_list(profile_id2):
            yield contacts_list_id1 + contacts_list_id2

        # Extract the profile IDs of both contact lists
        logger.debug("Extracting profile IDs.")
        contact_ids_id1 = [d['profile_id'] for d in contacts_list_id1]
        contact_ids_id2 = [d['profile_id'] for d in contacts_list_id2]

        # Make an intersection of the profile IDs
        logger.debug("Calculating intersection.")
        contact_set_id1 = set(contact_ids_id1[1:])
        contact_set_id2 = set(contact_ids_id2[1:])
        mutual_contacts = contact_set_id1.intersection(contact_set_id2)

        # The first two elements of the mutual contact list are the profiles themselves of which we are
        # retrieving mutual contacts
        mutual_contact_list = [contacts_list_id1[0], contacts_list_id2[0]]

        # For each ID in the intersection, obtain the original entry in the original dictionary
        # TODO: Understand how this really works
        logger.debug("Creating mutual contacts list.")

        for contact in mutual_contacts:
            entry = (item for item in contacts_list_id1 if item['profile_id'] == contact).next()
            mutual_contact_list.append(entry)

        logger.debug("Returning mutual contact list.")

        yield mutual_contact_list

    def _extract_login_data(self, html):
        pattern = re.compile("\"ACCOUNT_ID\":\"(\d+?)\",\"NAME\":\"(.+?)\"")
        match = pattern.search(html)
        if match:
            self.login_id = match.groups()[0]
            self.name = match.groups()[1]

        pattern = re.compile("\"startTime\":(\d+?),")
        match = pattern.search(html)
        if match:
            self.start_time = match.groups()[0]


class Rolodex(object):
    def __init__(self, browser, login_id, start_time, profile_name):
        logger.debug("Creating rolodex.")

        self.name = "name N\A"
        self.profile_name = profile_name
        self.login_id = login_id
        self.browser = browser
        self.start_time = start_time
        self.base_url = 'https://www.facebook.com/{0}'.format(profile_name)
        self.contact_url = self.base_url + '/friends'
        self._page_number = 1

        logger.debug("Rolodex created.")

    def num_entries(self):
        num_entries = 1

        logger.debug("Retrieving number of entries in rolodex.")

        html = self.browser.open(self.base_url).read()

        soup = BeautifulSoup(html, "html5lib")

        logger.debug("Trying to find all <span> elements in the retrieved HTML.")

        selection = soup.find_all('code')
        for nr, item in enumerate(selection):
            logger.debug("Analyzing <span> element #{0}".format(nr))

            if len(item.contents) > 0:
                logger.debug("The item contents has length > 0. Attempting to pattern match.")

                comment = item.contents[0]

                # logger.debug("Contents of item.contents[0] == {0}".format(comment))

                results = BeautifulSoup(comment, "html5lib").find_all('a', {"class": "_39g5",
                                                                            "href": re.compile(
                                                                                'http.+facebook\.com/.+friends.*')})

                logger.debug("Results came in...")
                logger.debug("Length of results: {0}".format(len(results)))

                for result_nr, elem in enumerate(results):
                    logger.debug("Match! Processing result #{0}".format(result_nr))
                    num_entries = re.search(r'(\d+)', elem.contents[0]).group(1)
                    return int(num_entries)

        return int(num_entries)

    def __iter__(self):
        return self

    def next(self):
        if self._page_number == 1:
            logger.debug("Opening browser to get first page.")

            page = self.browser.open(self.contact_url).read()

            logger.debug("First page retrieved (HTML).")

            self.contact_url = self._compose_url_from_html(page)

            logger.debug("Attempt to construct contact URL finished.")
            logger.debug("Attempting to extract contacts from HTML.")

            contacts = self._extract_contacts_from_html(page)

            logger.debug("Attempt to extract contacts. Moving on.")
        elif self._page_number > 1 and self.contact_url is None:
            logger.debug("The contact URL is None. Stop iterations.")
            raise StopIteration
        else:
            logger.debug("Opening browser to get next page.")

            page = self.browser.open(self.contact_url).read().decode('unicode_escape')

            logger.debug("Next page retrieved (Javascript).")

            self.contact_url = self._compose_url_from_script(page)

            logger.debug("Attempt to construct contact URL finished.")
            logger.debug("Attempting to extract contacts from Javascript.")

            contacts = self._extract_contacts_from_script(page)

            logger.debug("Next page retrieved.")

        self._page_number += 1
        return contacts

    def _compose_url_from_html(self, html):
        soup = BeautifulSoup(html, "html5lib")

        logger.debug("Trying to find all <script> elements in the retrieved HTML.")

        selection = soup.find_all('script')
        for nr, script in enumerate(selection):
            logger.debug("Analyzing <script> element #{0}".format(nr))

            if len(script.contents) > 0:
                logger.debug("The script has length > 0. Attempting to pattern match.")

                pattern2 = re.compile("pagelet_token:\"(.*)\",tab_key:\"friends\",lst:\"(\d+?):(\d+?):(\d+?)\"")
                match2 = pattern2.search(script.contents[0])
                if match2:
                    lst = []
                    self.pagelet_token = match2.groups()[0]
                    lst.append(match2.groups()[1])
                    lst.append(match2.groups()[2])
                    lst.append(match2.groups()[3])

                    continue

                pattern = re.compile("TimelineAppCollection\",\"enableContentLoader\",.*\"pagelet_timeline_app_collection_(.*?)\".*?\},\"(.*?)\"")
                match = pattern.search(script.contents[0])
                if match:
                    logger.debug("The pattern matched: match.groups()[0] == {0}".format(match.groups()[0]))
                    logger.debug("Continuing to extract elements.")

                    profile_id, research_id, number = match.groups()[0].split(':')
                    cursor = match.groups()[1]

                    logger.debug(
                        "profile_id == {0}, research_id == {1}, number == {2}".format(profile_id, research_id, number))
                    logger.debug("Breaking out of loop")

                    break
        else:
            logger.debug("Finished go through script without finding matches.")
            logger.info("This profile has no viewable friends.")
            return None

        composed_url = \
            'https://www.facebook.com/ajax/pagelet/generic.php/' \
            'AllFriendsAppCollectionPagelet?' \
            'dpr=1&data=%7B%22collection_token' \
            '%22%3A%22' \
            '{profile_id}' \
            '%3A' \
            '{research_id}' \
            '%3A' \
            '{number}' \
            '%22%2C%22' \
            'cursor' \
            '%22%3A%22' \
            '{cursor}' \
            '%3D%22%2C%22' \
            'disablepager' \
            '%22%3A' \
            'false' \
            '%2C%22' \
            'overview' \
            '%22%3A' \
            'false' \
            '%2C%22' \
            'profile_id' \
            '%22%3A%22' \
            '{profile_id}' \
            '%22%2C%22' \
            'pagelet_token' \
            '%22%3A%22' \
            '{pagelet_token}' \
            '%22%2C%22' \
            'tab_key' \
            '%22%3A%22' \
            'friends' \
            '%22%2C%22' \
            'lst' \
            '%22%3A%22' \
            '{login_id}' \
            '%3A' \
            '{profile_id}' \
            '%3A' \
            '{start_time}' \
            '%22%2C%22ftid%22%3Anull%2C%22order%22%3Anull%2C%22sk%22%3A%22' \
            'friends%22%2C%22importer_state%22%3Anull%7D&' \
            '__user={login_id}' \
            '&__a=&__dyn=&__af=i0&__req=&__be=-1&__pc=PHASED%3ADEFAULT&__rev=' \
                .format(profile_id=profile_id, research_id=research_id, number=number, cursor=cursor,
                        login_id=self.login_id, start_time=self.start_time, pagelet_token=self.pagelet_token)

        logger.debug("URL for next iteration composed.")

        return composed_url

    def _compose_url_from_script(self, script):  # DRY !
        logger.debug("Trying to match pattern in the retrieved Javascript.")

        pattern = re.compile("TimelineAppCollection\",\"enableContentLoader\",.*\"pagelet_timeline_app_collection_(.*?)\".*?\},\"(.*?)\"")
        match = pattern.search(script)
        if match:
            logger.debug("The pattern matched: match.groups()[0] == {0}".format(match.groups()[0]))
            logger.debug("Continuing to extract elements.")

            profile_id, research_id, number = match.groups()[0].split(':')
            cursor = match.groups()[1]

            logger.debug("profile_id == {0}, research_id == {1}, number == {2}".format(profile_id, research_id, number))
        else:
            return None

        composed_url = \
            'https://www.facebook.com/ajax/pagelet/generic.php/' \
            'AllFriendsAppCollectionPagelet?' \
            'dpr=1&data=%7B%22collection_token' \
            '%22%3A%22' \
            '{profile_id}' \
            '%3A' \
            '{research_id}' \
            '%3A' \
            '{number}' \
            '%22%2C%22' \
            'cursor' \
            '%22%3A%22' \
            '{cursor}' \
            '%3D%22%2C%22' \
            'disablepager' \
            '%22%3A' \
            'false' \
            '%2C%22' \
            'overview' \
            '%22%3A' \
            'false' \
            '%2C%22' \
            'profile_id' \
            '%22%3A%22' \
            '{profile_id}' \
            '%22%2C%22' \
            'pagelet_token' \
            '%22%3A%22' \
            '{pagelet_token}' \
            '%22%2C%22' \
            'tab_key' \
            '%22%3A%22' \
            'friends' \
            '%22%2C%22' \
            'lst' \
            '%22%3A%22' \
            '{login_id}' \
            '%3A' \
            '{profile_id}' \
            '%3A' \
            '{start_time}' \
            '%22%2C%22ftid%22%3Anull%2C%22order%22%3Anull%2C%22sk%22%3A%22' \
            'friends%22%2C%22importer_state%22%3Anull%7D&' \
            '__user={login_id}' \
            '&__a=&__dyn=&__af=i0&__req=&__be=-1&__pc=PHASED%3ADEFAULT&__rev=' \
                .format(profile_id=profile_id, research_id=research_id, number=number, cursor=cursor,
                        login_id=self.login_id, start_time=self.start_time, pagelet_token=self.pagelet_token)

        logger.debug("URL for next iteration composed.")

        return composed_url

    def _extract_contacts_from_html(self, html):
        # Debug code for timeout bug TODO
        logger.debug("Attempting to extract contacts from HTML.")
        file_path = '/var/tmp/' + time.strftime("%Y%m%d-%H%M%S") + '_html.txt'
        logger.debug("File path: '{0}'.".format(file_path))

        with open(file_path, 'w') as html_file:
            html_file.write(html)
            logger.debug("Wrote HTML to file '{0}'.".format(file_path))

        contacts = []
        soup = BeautifulSoup(html, "html5lib")

        logger.debug("Trying to find all <title> elements in the retrieved HTML.")

        selection = soup.find_all('title')
        for nr, item in enumerate(selection):
            try:
                if item.attrs['id'] == "pageTitle":
                    self.name = item.contents[0]
            except:
                continue

        # First element of the contact list is the profile itself of which we are retrieving contacts
        contacts.append({"name": self.name, "uri": self.base_url, "profile_id": self.profile_name})

        logger.debug("Trying to find all <code> elements in the retrieved HTML.")

        selection = soup.find_all('code')
        for nr, item in enumerate(selection):
            logger.debug("Analyzing <code> element #{0}".format(nr))

            if len(item.contents) > 0:
                logger.debug("The item contents has length > 0. Attempting to pattern match.")

                comment = item.contents[0]

                # logger.debug("Contents of item.contents[0] == {0}".format(comment))

                results = BeautifulSoup(comment, "html5lib").find_all('a',
                                                                      {"data-hovercard-prefer-more-content-show": "1",
                                                                       "data-gt": re.compile('.*')}, )

                logger.debug("Results came in...")
                logger.debug("Length of results: {0}".format(len(results)))

                for result_nr, elem in enumerate(results):
                    logger.debug("Match! Processing result #{0}".format(result_nr))

                    link = elem.attrs['href'].replace("\\", "")
                    try:  # First try to see if the id is hidden in the url, like https://www.facebook.com/profile.php?id=100005592845863
                        profile_id = re.search(r'profile\.php\?id=(.*?)&.*hc_location=friends_tab', link).group(1)
                        contacts.append({"name": elem.contents[0], "uri": link, "profile_id": profile_id})
                        logger.debug("Profile_id '{0}' appended to contact list".format(profile_id))
                        continue
                    except AttributeError as e:
                        pass  # Fallback to the next try

                    try:  # Else the id is not a number but a string
                        profile_id = re.search(r'\.facebook\.com/(.*)\?.*hc_location=friends_tab', link).group(1)
                        contacts.append({"name": elem.contents[0], "uri": link, "profile_id": profile_id})
                        logger.debug("Profile_id '{0}' appended to contact list".format(profile_id))
                    except AttributeError as e:
                        # No profile found. The profile could be a follower not a friend.
                        pass

                        # logger.debug(
                        #     "Elements extracted: name == {0}, profile_id == {1}, uri == {2}".format(elem.contents[0],
                        #                                                                             profile_id, link))
        logger.debug("Returning contacts.")

        return contacts

    def _extract_contacts_from_script(self, script):
        # Debug code for timeout bug TODO
        logger.debug("Attempting to extract contacts from Javascript.")
        file_path = '/var/tmp/' + time.strftime("%Y%m%d-%H%M%S") + '_js.txt'
        logger.debug("File path: '{0}'.".format(file_path))

        with open(file_path, 'w') as js_file:
            # js_file.write(script)
            logger.debug("TODO Wrote HTML to file '{0}'.".format(file_path))

        logger.debug("Attempting to pattern match: 'payload\":\"(.*)\",\"jsmods\"'")

        contacts = []
        pattern = re.compile("payload\":\"(.*)\",\"jsmods\"")
        html = re.findall(pattern, script)
        if len(html) > 0:
            logger.debug("The item contents has length > 0. Attempting to pattern match more.")

            # Different than the other parsers, this needs to be a HTML parser, otherwise we get an exception
            results = BeautifulSoup(html[0], "html.parser").find_all('a',
                                                                     {"data-hovercard-prefer-more-content-show": "1",
                                                                      "data-gt": re.compile('.*')})
            for result_nr, elem in enumerate(results):
                logger.debug("Match! Processing result #{0}".format(result_nr))

                # TODO: This code is not specific enough. When there are followers, the code below
                # will also succeed, even though it should not. For now the function _extract_contacts_from_html
                # takes care of never even ending up here, because with only followers a contact_url for the
                # next page (i.e. ending up here) cannot be formed. But it is not water tight.
                link = elem.attrs['href'].replace("\\", "")
                try:  # First try to see if the id is hidden in the url, like https://www.facebook.com/profile.php?id=100005592845863
                    profile_id = re.search(r'profile\.php\?id=(.*?)&', link).group(1)
                except AttributeError as e:  # Else the id is not a number but a string
                    profile_id = re.search(r'\.facebook\.com/(.*)\?', link).group(1)

                # logger.debug(
                #     "Elements extracted: name == {0}, profile_id == {1}, uri == {2}".format(elem.contents[0],
                #                                                                             profile_id, link))
                logger.debug("profile_id: {0}".format(profile_id))

                try:
                    # Remove strange <\/a><\/div> like string from the end. This is due to the HTML parser.
                    name = re.sub('<.*$', '', elem.contents[0])
                    # logger.debug("name: {0}".format(name))
                    contacts.append({"name": name, "uri": link, "profile_id": profile_id})
                except Exception as e:
                    # pdb.set_trace()
                    pass

                logger.debug("Appended to contact list.")

        logger.debug("Returning contacts.")

        return contacts
